#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import io
import json
import math
import re
import sys
import zipfile
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

import numpy as np
from openpyxl import load_workbook


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def load_dev_module(path: Path):
    spec = importlib.util.spec_from_file_location("v019_frozen_shared", path)
    if spec is None or spec.loader is None:
        raise RuntimeError("cannot load shared frozen module")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def norm_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


SYNONYMS = {
    "j": ("j", "advanceratio", "advancecoefficient"),
    "ct": ("ct", "thrustcoefficient", "coefficientofthrust"),
    "cp": ("cp", "powercoefficient", "coefficientofpower"),
    "rpm": ("rpm", "rotationspeed", "rotationalspeed", "motorspeed", "propspeed"),
    "airspeed": ("airspeed", "velocity", "freestreamvelocity", "vinf", "v∞", "windvelocity"),
    "thrust": ("thrust", "axialforce", "forcez", "fz"),
    "torque": ("torque", "momentz", "mz", "q"),
    "power": ("power", "shaftpower", "mechanicalpower"),
}


def header_role(header: str) -> str | None:
    h = norm_header(header)
    for role, options in SYNONYMS.items():
        if h in options:
            return role
        if len(h) >= 3 and any(opt in h for opt in options if len(opt) >= 3):
            return role
    return None


def numeric(v: object) -> float | None:
    if v is None or isinstance(v, bool):
        return None
    try:
        x = float(v)
    except (TypeError, ValueError):
        return None
    return x if math.isfinite(x) else None


def workbook_tables(data: bytes, source: str) -> Iterable[Tuple[str, List[List[object]]]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        yield f"{source}::{ws.title}", rows


def delimited_tables(data: bytes, source: str) -> Iterable[Tuple[str, List[List[object]]]]:
    text = data.decode("utf-8", "replace")
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return
    for delimiter in (",", ";", "\t", None):
        rows: List[List[object]] = []
        try:
            if delimiter is None:
                rows = [re.split(r"\s+", line.strip()) for line in lines]
            else:
                rows = list(csv.reader(lines, delimiter=delimiter))
        except Exception:
            continue
        if rows and max(len(r) for r in rows) >= 3:
            yield source, rows
            return


def tables_from_member(name: str, data: bytes) -> Iterable[Tuple[str, List[List[object]]]]:
    suffix = Path(name).suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        yield from workbook_tables(data, name)
    elif suffix in (".csv", ".txt", ".dat", ".tsv"):
        yield from delimited_tables(data, name)


def detect_table(rows: Sequence[Sequence[object]]) -> Tuple[int, Dict[str, int]] | None:
    best = None
    for i, row in enumerate(rows[:40]):
        mapping: Dict[str, int] = {}
        for j, cell in enumerate(row):
            role = header_role(str(cell or ""))
            if role is not None and role not in mapping:
                mapping[role] = j
        direct = all(k in mapping for k in ("j", "ct", "cp"))
        raw = all(k in mapping for k in ("rpm", "airspeed", "thrust")) and ("power" in mapping or "torque" in mapping)
        score = len(mapping) + (10 if direct else 0) + (8 if raw else 0)
        if (direct or raw) and (best is None or score > best[0]):
            best = (score, i, mapping)
    return None if best is None else (best[1], best[2])


def unit_scale(header: object, role: str) -> float:
    h = str(header or "").lower()
    if role == "airspeed":
        if "km/h" in h or "kmh" in h: return 1.0 / 3.6
        if "mph" in h: return 0.44704
    if role == "thrust":
        if "lbf" in h or "lb_f" in h: return 4.4482216153
        if "kgf" in h: return 9.80665
        if "gf" in h or "gram" in h: return 0.00980665
        if "mn" in h and "nm" not in h: return 0.001
    if role == "torque":
        if "mn m" in h or "mnm" in norm_header(h): return 0.001
        if "n mm" in h or "nmm" in norm_header(h): return 0.001
        if "gcm" in norm_header(h): return 9.80665e-5
    if role == "power" and "kw" in h:
        return 1000.0
    return 1.0


def geometry_from_path(path: str, catalog: Sequence[Dict[str, object]]) -> Dict[str, object] | None:
    lower = path.lower()
    ranked = []
    for row in catalog:
        key = str(row.get("key", "")).lower()
        aliases = [str(x).lower() for x in row.get("aliases", [])]
        hits = [x for x in [key, *aliases] if x and x in lower]
        if hits:
            ranked.append((max(len(x) for x in hits), row))
    if ranked:
        return max(ranked, key=lambda x: x[0])[1]
    m = re.search(r"(?<!\d)(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)(?!\d)", path)
    if m:
        return {"key": Path(path).stem, "diameter_in": float(m.group(1)), "pitch_in": float(m.group(2)), "aliases": []}
    return None


def rpm_from_path(path: str) -> float | None:
    patterns = (r"(?:rpm[_ -]?)(\d{3,6})", r"[_ -](\d{3,6})(?:rpm)?(?:\.[a-z0-9]+)?$")
    for pattern in patterns:
        m = re.search(pattern, path.lower())
        if m:
            value = float(m.group(1))
            if 100 <= value <= 200000:
                return value
    return None


def rows_to_observations(
    table_name: str,
    rows: Sequence[Sequence[object]],
    geometry: Dict[str, object],
    rho: float,
) -> List[Dict[str, float]]:
    detected = detect_table(rows)
    if detected is None:
        return []
    header_idx, mapping = detected
    headers = list(rows[header_idx])
    D_m = float(geometry["diameter_in"]) * 0.0254
    fixed_rpm = rpm_from_path(table_name)
    observations = []
    for row in rows[header_idx + 1 :]:
        def value(role: str) -> float | None:
            if role not in mapping or mapping[role] >= len(row): return None
            x = numeric(row[mapping[role]])
            if x is None: return None
            return x * unit_scale(headers[mapping[role]], role)
        j = value("j"); ct = value("ct"); cp = value("cp")
        rpm = value("rpm")
        if rpm is None: rpm = fixed_rpm
        if j is not None and ct is not None and cp is not None and rpm is not None:
            pass
        else:
            airspeed = value("airspeed"); thrust = value("thrust"); power = value("power"); torque = value("torque")
            if rpm is None or airspeed is None or thrust is None or (power is None and torque is None):
                continue
            n = rpm / 60.0
            if n <= 0 or D_m <= 0: continue
            if power is None: power = 2.0 * math.pi * n * float(torque)
            j = airspeed / (n * D_m)
            ct = thrust / (rho * n * n * D_m ** 4)
            cp = power / (rho * n ** 3 * D_m ** 5)
        if not all(math.isfinite(float(x)) for x in (j, ct, cp, rpm)):
            continue
        if not (-0.35 <= float(j) <= 3.0 and 0 < float(rpm) <= 200000 and abs(float(ct)) <= 3 and abs(float(cp)) <= 3):
            continue
        observations.append({"j": float(j), "ct": float(ct), "cp": float(cp), "rpm": float(rpm)})
    return observations


def build_families(module, archive: Path, protocol: Dict[str, object]):
    by: Dict[str, Dict[str, object]] = {}
    candidates = set(protocol["holdout"]["candidate_member_paths"])
    catalog = protocol["holdout"]["geometry_catalog"]
    rho = float(protocol["physics"]["air_density_kg_m3"])
    invalid_members = []
    with zipfile.ZipFile(archive) as z:
        for name in sorted(candidates):
            if name not in z.namelist():
                invalid_members.append({"path": name, "reason": "missing_from_verified_archive"})
                continue
            geometry = geometry_from_path(name, catalog)
            if geometry is None:
                invalid_members.append({"path": name, "reason": "no_frozen_geometry_match"})
                continue
            data = z.read(name)
            found = []
            try:
                for table_name, rows in tables_from_member(name, data):
                    found.extend(rows_to_observations(table_name, rows, geometry, rho))
            except Exception as exc:
                invalid_members.append({"path": name, "reason": f"parse_exception:{type(exc).__name__}"})
                continue
            if not found:
                invalid_members.append({"path": name, "reason": "no_supported_performance_table"})
                continue
            key = str(geometry["key"])
            slot = by.setdefault(key, {"geometry": geometry, "rows": [], "members": []})
            slot["rows"].extend(found)
            slot["members"].append(name)
    families = []
    invalid_families = []
    for key, slot in sorted(by.items()):
        rows = slot["rows"]
        # Exact duplicates can arise from repeated export sheets.
        unique = sorted({(round(x["j"], 12), round(x["rpm"], 8), round(x["ct"], 12), round(x["cp"], 12)) for x in rows})
        if len(unique) < int(protocol["validity"]["minimum_rows"]):
            invalid_families.append({"family": key, "reason": "too_few_rows", "rows": len(unique)})
            continue
        arr = np.asarray(unique, float)
        if np.ptp(arr[:, 2]) < protocol["validity"]["minimum_output_range"] or np.ptp(arr[:, 3]) < protocol["validity"]["minimum_output_range"]:
            invalid_families.append({"family": key, "reason": "degenerate_output_range", "rows": len(unique)})
            continue
        g = slot["geometry"]
        families.append(module.Family(
            name=key,
            diameter_in=float(g["diameter_in"]),
            pitch_in=float(g["pitch_in"]),
            j=arr[:, 0], rpm=arr[:, 1], ct=arr[:, 2], cp=arr[:, 3],
        ))
    return families, invalid_members, invalid_families


def atlas_from_json(module, model: Dict[str, object]):
    a = model["atlas"]
    return module.Atlas(
        bounds={k: float(v) for k, v in a["bounds"].items()},
        coeffs={k: np.asarray(v, float) for k, v in a["coefficients"].items()},
        tau_prior_beta=np.asarray(a["tau_prior_beta"], float),
        tau_lo=float(a["tau_lo"]), tau_hi=float(a["tau_hi"]),
        output_floors={k: float(v) for k, v in a["output_floors"].items()},
        prior_lambda=float(a["prior_lambda"]),
    )


def aggregate(rows: Sequence[Dict[str, object]], method: str) -> Dict[str, float]:
    v = np.asarray([float(x[f"{method}_joint"]) for x in rows])
    fixed = np.asarray([float(x["fixed_joint"]) for x in rows])
    ratio = v / np.maximum(fixed, 1e-12)
    return {
        "median_joint_nrmse": float(np.median(v)),
        "geometric_mean_joint_nrmse": float(np.exp(np.mean(np.log(np.maximum(v, 1e-12))))),
        "median_ratio_vs_fixed": float(np.median(ratio)),
        "geometric_mean_ratio_vs_fixed": float(np.exp(np.mean(np.log(np.maximum(ratio, 1e-12))))),
        "win_fraction_vs_fixed": float(np.mean(ratio < 1.0)),
        "worst_ratio_vs_fixed": float(np.max(ratio)),
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--protocol", required=True)
    ap.add_argument("--model", required=True)
    ap.add_argument("--shared-module", required=True)
    ap.add_argument("--archive", required=True)
    ap.add_argument("--output-dir", required=True)
    ap.add_argument("--self-test", action="store_true")
    args = ap.parse_args()
    if args.self_test:
        assert norm_header("C_T [-]") == "ct"
        assert abs(unit_scale("Velocity [km/h]", "airspeed") - 1/3.6) < 1e-12
        print("SELF_TEST_PASS")
        return 0
    protocol_path=Path(args.protocol); model_path=Path(args.model); module_path=Path(args.shared_module); archive_path=Path(args.archive)
    protocol=json.loads(protocol_path.read_text()); model=json.loads(model_path.read_text())
    for path,key in ((protocol_path,"protocol_sha256"),(model_path,"model_sha256"),(module_path,"shared_module_sha256")):
        expected=protocol["integrity"].get(key)
        if expected and sha256_bytes(path.read_bytes()) != expected: raise RuntimeError(f"integrity mismatch: {key}")
    if sha256_bytes(archive_path.read_bytes()) != protocol["holdout"]["archive_sha256"]: raise RuntimeError("ENOLA archive SHA-256 mismatch")
    module=load_dev_module(module_path); atlas=atlas_from_json(module,model)
    families,invalid_members,invalid_families=build_families(module,archive_path,protocol)
    records=[]
    for f in families:
        probes=module.select_probes(f,int(protocol["binding"]["probe_count"]))
        scored=np.asarray([i for i in range(len(f.j)) if i not in set(probes.tolist())],int)
        if len(scored)<protocol["validity"]["minimum_scored_rows"]:
            invalid_families.append({"family":f.name,"reason":"too_few_scored_rows","rows":len(f.j)})
            continue
        safe,tau,alpha=module.safe_atlas_predict(atlas,f,probes)
        raw,_=module.atlas_predict(atlas,f,probes)
        fixed=module.fixed_cubic_predict(f,probes,"fixed")
        plain=module.fixed_cubic_predict(f,probes,"j")
        row={"family":f.name,"diameter_in":f.diameter_in,"pitch_in":f.pitch_in,"rows":len(f.j),"probe_indices":";".join(map(str,probes.tolist())),"atlas_tau":tau,"atlas_weight":alpha}
        for name,pred in (("safe_atlas",safe),("raw_atlas",raw),("fixed",fixed),("plain_j",plain)):
            joint,errs=module.family_joint_error(f,pred,scored,atlas.output_floors)
            row[f"{name}_joint"]=joint;row[f"{name}_ct"]=errs["ct"];row[f"{name}_cp"]=errs["cp"]
        records.append(row)
    methods={name:aggregate(records,name) for name in ("safe_atlas","raw_atlas","fixed","plain_j")} if records else {}
    safe=methods.get("safe_atlas",{})
    plain=np.asarray([float(x["plain_j_joint"]) for x in records]) if records else np.asarray([])
    cand=np.asarray([float(x["safe_atlas_joint"]) for x in records]) if records else np.asarray([])
    plain_ratio=float(np.exp(np.mean(np.log(np.maximum(cand/np.maximum(plain,1e-12),1e-12))))) if len(records) else float("inf")
    plain_win=float(np.mean(cand<plain)) if len(records) else 0.0
    gates={
      "minimum_valid_families":{"value":len(records),"threshold":protocol["gates"]["minimum_valid_families"],"pass":len(records)>=protocol["gates"]["minimum_valid_families"]},
      "zero_catastrophic":{"value":sum(not math.isfinite(float(x["safe_atlas_joint"])) or float(x["safe_atlas_joint"])>=protocol["gates"]["catastrophic_joint_nrmse"] for x in records),"threshold":0,"pass":all(math.isfinite(float(x["safe_atlas_joint"])) and float(x["safe_atlas_joint"])<protocol["gates"]["catastrophic_joint_nrmse"] for x in records)},
      "median_joint_nrmse":{"value":safe.get("median_joint_nrmse",float("inf")),"threshold":protocol["gates"]["median_joint_nrmse_lte"],"pass":safe.get("median_joint_nrmse",float("inf"))<=protocol["gates"]["median_joint_nrmse_lte"]},
      "gm_ratio_vs_fixed":{"value":safe.get("geometric_mean_ratio_vs_fixed",float("inf")),"threshold":protocol["gates"]["geometric_mean_ratio_vs_fixed_lte"],"pass":safe.get("geometric_mean_ratio_vs_fixed",float("inf"))<=protocol["gates"]["geometric_mean_ratio_vs_fixed_lte"]},
      "win_fraction_vs_fixed":{"value":safe.get("win_fraction_vs_fixed",0.0),"threshold":protocol["gates"]["win_fraction_vs_fixed_gte"],"pass":safe.get("win_fraction_vs_fixed",0.0)>=protocol["gates"]["win_fraction_vs_fixed_gte"]},
      "gm_ratio_vs_plain_j":{"value":plain_ratio,"threshold":protocol["gates"]["geometric_mean_ratio_vs_plain_j_lte"],"pass":plain_ratio<=protocol["gates"]["geometric_mean_ratio_vs_plain_j_lte"]},
      "win_fraction_vs_plain_j":{"value":plain_win,"threshold":protocol["gates"]["win_fraction_vs_plain_j_gte"],"pass":plain_win>=protocol["gates"]["win_fraction_vs_plain_j_gte"]},
    }
    result={"protocol_id":protocol["protocol_id"],"primary_pass":all(g["pass"] for g in gates.values()),"summary":{"valid_family_count":len(records),"invalid_family_count":len(invalid_families),"invalid_member_count":len(invalid_members),"methods":methods,"safe_vs_plain_j":{"geometric_mean_ratio":plain_ratio,"win_fraction":plain_win},"median_tau":float(np.median([x["atlas_tau"] for x in records])) if records else None,"median_atlas_weight":float(np.median([x["atlas_weight"] for x in records])) if records else None},"gates":gates,"invalid_members":invalid_members,"invalid_families":invalid_families}
    out=Path(args.output_dir);out.mkdir(parents=True,exist_ok=True)
    (out/"results.json").write_text(json.dumps(result,indent=2,sort_keys=True)+"\n")
    if records:
        with (out/"per_family.csv").open("w",newline="") as fh:
            w=csv.DictWriter(fh,fieldnames=list(records[0]));w.writeheader();w.writerows(records)
    verdict=["# Causal Loom v0.19 frozen ENOLA result","",f"Primary pass: **{result['primary_pass']}**",f"Valid configurations: **{len(records)}**",f"Invalid configurations: **{len(invalid_families)}**",""]
    for key,g in gates.items(): verdict.append(f"- {key}: {g['value']} (gate {g['threshold']}) — {'PASS' if g['pass'] else 'FAIL'}")
    (out/"verdict.md").write_text("\n".join(verdict)+"\n")
    print("V019_FROZEN_RESULT="+json.dumps({"primary_pass":result["primary_pass"],"summary":result["summary"],"gates":gates},sort_keys=True))
    return 0 if result["primary_pass"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
